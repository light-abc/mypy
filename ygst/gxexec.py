#!/usr/bin/env python
# coding:utf-8

# import xlsxwriter
import time
import os
import re
import string
from openpyxl import load_workbook

os.chdir(r'C:\Users\168\Documents')
date=time.strftime('%m%d')
name=r"C:\Users\168\Desktop\巡检\巡检记录--9点 - 1201.xlsx"

wb = load_workbook(name)
ws = wb.active

n=0
for i in range(2, 87):
    n = n + 1
 #   val = re.sub(r'%', "", ws['D' + str(i)].value)
    val1 = ws['D' + str(i)].value
    ws.cell(row=n + 1, column=4, value=val1)

    val2 = ws['E' + str(i)].value
    ws.cell(row=n + 1, column=5, value=val2)

    val3 = ws['F' + str(i)].value
    ws.cell(row=n + 1, column=6, value=val3)

#wb.save(name)

for j in ws.rows:
    for n in j:
        print(n.value, end="\t")   # n.value 获取单元格的值
    print()

    # sheets = wb.sheetnames
    # sheet_first = sheets[0]
    # ws = wb[sheet_first]
    # rows = ws.rows
    # columns = ws.columns

    # for row in rows:
    #     line = [col.value for col in row]
    #     print(line)

    #sheet = wb['cpu平均利用率','内存平均利用率','磁盘使用情况']
    # print(ws["D:F"])
    # for column in ws["D:F"]:
    #     for cell in column:
    #         print(cell.value)
    # ws["D2"].number_format = openpyxl.styles.numbers.FORMAT_NUMBER