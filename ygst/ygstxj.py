#!/usr/bin/env python
# coding:utf-8

import shutil
# import string
import time
import os
from openpyxl import Workbook
# from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, colors, Alignment, Border, Side, numbers

date=time.strftime('%m%d')
os.chdir(r'C:\Users\168\Documents')

title=['ip地址','服务器名称','主机名','cpu平均利用率','内存平均利用率','磁盘使用情况','网络连接','日志检查','系统漏洞',\
       '软件运行情况','异常问题记录','巡检记录人']
fwqm=["Nginx反向代理","Nginx反向代理","前端服务器","前端服务器","前端服务器","动态网关服务器","动态网关服务器",\
      "动态网关服务器","动态网关服务器","动态网关服务器","动态网关服务器","动态网关服务器","动态网关服务器","动态网关服务器","动态网关服务器","动态网关服务器","动态网关服务器",\
      "微服务服务器","微服务服务器","微服务服务器","微服务服务器","微服务服务器","微服务服务器",\
      "跑缓存线程","应用数据库","应用数据库","应用数据库","应用数据库","应用数据库","应用数据库","应用数据库","应用数据库","应用数据库","应用数据库","应用数据库","应用数据库",\
      "前端服务器","前端服务器",\
      "文件服务器","文件服务器","文件服务器","文件服务器","文件服务器","文件服务器","文件服务器","文件服务器","文件服务器","文件服务器","文件服务器","文件服务器","文件服务器","文件服务器",\
      "缓存服务器","缓存服务器","缓存服务器","缓存服务器",\
      "日志Mycat服务器","日志Mycat服务器","应用Mycat服务器","应用Mycat服务器","消息队列服务器",\
      "注册中心","定时任务中心","注册中心",\
      "应用数据库","应用数据库","应用数据库","应用数据库","应用数据库","应用数据库","应用数据库","应用数据库","应用数据库","应用数据库","应用数据库","应用数据库","应用数据库","应用数据库","应用数据库","应用数据库",\
      "微服务服务器","微服务服务器","监控服务器",\
      "日志数据库","日志数据库"]
hostn=["Nginx1","Nginx2","Back-C1","Back-C2","Back-C3","GW1","GW2","GW3","GW4","GW5","GW6","GW7","GW8","GW9","GW10","GW11","GW12",\
       "MS1","MS2","MS3","MS4","MS5","MS6","finance-task","DB1","DB2","DB3","DB4","DB5","DB6","DS1","DS2","DS3","DS4","DS5","DS6",\
       "Back-C4","Back-C5",\
       "File1","File2","File3","File4","File5","File6","File7","File8","File9","File10","File11","File12","File13","File14",\
       "RD1","RD2","RD3","RD4",\
       "Log-Mycat1","Log-Mycat2","DB-Mycat1","DB-Mycat2","activemq",\
       "eureka","task","eureka2",\
       "DSS1","DSS2","DSS3","DSS4","DSS5","DSS6","DSSS1","DSSS2","DSSS3","DSSS4","DSSS5","DSSS6","DB7","DS7","DSS7","DSSS7",\
       "MS7","MS8","zabbix","LDB1","LDB2"]
val1='无异常'
val2='无'
mz='黄亮亮'

t = time.strftime('%H')
if t <= '12':
    t = '9'
elif t > '12':
    t = '17'
name = '巡检记录--' + t + '点 - ' + date + '.xlsx'

wb = Workbook()

ws = wb.active

ws.append(title)

n = 0
for i in range(160,224):
    n += 1
    ws.cell(row=n + 1, column=1, value='10.0.18.'+str(i))
for i in range(227,248):
    n += 1
    ws.cell(row=n + 1, column=1, value='10.0.18.' + str(i))

for i,m in enumerate(fwqm):
    i1 = i + 1
    ws.cell(row=i1 + 1, column=2, value=fwqm[i])
for i,m in enumerate(hostn):
    i2 = i + 1
    ws.cell(row=i2 + 1, column=3, value=hostn[i])

with open('xunjian.txt', 'r+', encoding='utf-8') as f:
    s1 = f.readlines()
f.close()

for i,s in enumerate(s1):
    i += 1
    list = s.split()
    # if s[1] is str:
    cpu = float(list[0])/100
    mem = float(list[1])/100
    disk = float(list[2])/100
    # '{:.2f}%'.format(cpu)
    ws.cell(row=i+1, column=4, value=cpu)
    ws.cell(row=i+1, column=5, value=mem)
    ws.cell(row=i+1, column=6, value=disk)
    ws.cell(row=i+1, column=7, value=val1)
    ws.cell(row=i+1, column=8, value=val1)
    ws.cell(row=i+1, column=9, value=val1)
    ws.cell(row=i+1, column=10, value=val1)
    ws.cell(row=i+1, column=11, value=val2)
    ws.cell(row=i+1, column=12, value=mz)
# ws.sheet_properties.tabColor = "1072BA"

border = Border(left=Side(border_style='thin',color='000000'),
right=Side(border_style='thin',color='000000'),
top=Side(border_style='thin',color='000000'),
bottom=Side(border_style='thin',color='000000'))

for z in 'ABCDEFGHIJKL':
    for i in range(1, 87):
        ws[str(z)+str(i)].font = Font(name='宋体', size=11, italic=False, color=colors.BLACK, bold=False)
        ws[str(z)+str(i)].alignment = Alignment(horizontal='center', vertical='center')
        ws[str(z)+str(i)].number_format=numbers.BUILTIN_FORMATS[10]
        ws.column_dimensions[str(z)].width = 16
        ws.row_dimensions[i].height = 15
        ws[str(z)+str(i)].border = border

# for z in list(string.ascii_uppercase):
# row_max = ws.max_row
# con_max = ws.max_column

# for j in ws.rows:
#     for n in j:
#         print(n.value, end="\t")   # n.value 获取单元格的值
#     print()

wb.save(name)
