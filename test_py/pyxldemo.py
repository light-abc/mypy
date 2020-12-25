#!/usr/bin/env python
import openpyxl
from openpyxl.styles import numbers
import os

# from openpyxl.utils import get_column_letter
# from openpyxl.utils import column_index_from_string
#
# print(get_column_letter)
# print(column_index_from_string)

os.chdir(r'C:\Users\168\Documents')
#wb = Workbook()
wb2 = openpyxl.load_workbook('巡检记录--9点 - 1203.xlsx')
print(wb2.sheetnames)

ws = wb2.active

a5 = ws['D5']
print('(%s,%s) is %s.' % (a5.column, a5.row, a5.value))

# a_sheet = wb2['巡检记录']
# print(a_sheet.title)
number_format = 'General'
# print(ws.max_column)
# print(ws.max_row)
# Data can be assigned directly to cells
row_range = ws['D2':'D86']
for row in row_range:
    for cell in row:
        print(cell.value)
        # a=eval(cell.value)
        # print(type(a))
        print(cell.number_format)
        # ws['D'] = a
        # cell.number_format = numbers.FORMAT_NUMBER
# for cell in ws['D']:
#     print(cell.value)
    #ws['D'].number_format = numbers.FORMAT_NUMBER
#print(ws['D5'].number_format)
#ws.cell['D2'].number_format = 'General'
#ws['D2':'D86'].style = 'Percent'
# Rows can also be appended
#ws.append([1, 2, 3])

# Python types will automatically be converted
#ws['A2'] = datetime.datetime.now()

# Save the file
wb2.save("巡检记录--9点 - 1203.xlsx")